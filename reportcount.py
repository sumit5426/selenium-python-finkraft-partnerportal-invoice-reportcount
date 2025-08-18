import time
import os
import uuid
import glob
import tempfile
import psycopg2
import pytz
from psycopg2 import sql
from pymongo import MongoClient
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.webdriver.support.wait import WebDriverWait
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from dotenv import load_dotenv
from dateutil import parser
from datetime import datetime
import openpyxl

load_dotenv()
with open("run_id.txt", "r") as f:
    run_id = f.read().strip()

morning_workspaces = [
    {
        "uri": "https://cleartrip.finkraft.ai/auth/signin",
        "workspace_name": "Unitus capital",
        "table_name": "flight_recon_main",
        "db_workspace_name": "Unitus capital"
    },

    {"uri": "https://balmer.finkraft.ai/auth/signin",
     "workspace_name": "Enforcement directorate",
     "table_name": "airline_recon_balmer",
     "db_workspace_name": "ENFORCEMENT DIRECTORATE"
     },
    {
        "uri": "https://myyatra.finkraft.ai/auth/signin",
        "workspace_name": "Minda industries limited yatra",
        "table_name": "flight_recon_yatra",
        "db_workspace_name": "MINDA INDUSTRIES LIMITED Yatra"
    },
    {
        "uri": "https://mmt.finkraft.ai/auth/signin",
        "workspace_name": "Mankind pharma",
        "table_name": "airline_recon_mmt",
        "db_workspace_name": "MANKIND PHARMA LIMITED"
    },
    {
        "uri": "https://pyt.finkraft.ai/auth/signin",
        "workspace_name": "Om sai intex",
        "table_name": "airline_recon_py",
        "db_workspace_name": "Om Sai Intex"
    },
    {
        "uri": "https://fcm.finkraft.ai/auth/signin",
        "workspace_name": "Infifresh",
        "table_name": "airline_recon_fcm",
        "db_workspace_name": "INFIFRESH"
    },
    {
        "uri": "https://bcd.finkraft.ai/auth/signin",
        "workspace_name": "Cp kelco",
        "table_name": "airline_recon_bcd",
        "db_workspace_name": "CP Kelco"
    },
    {
        "uri": "https://atpi.finkraft.ai/auth/signin",
        "workspace_name": "British international investment",
        "table_name": "airline_recon_atpi",
        "db_workspace_name": "British International Investment"
    }
]

afternoon_workspaces = [
    {
        "uri": "https://cleartrip.finkraft.ai/auth/signin",
        "workspace_name": "Saksoft",
        "table_name": "flight_recon_main",
        "db_workspace_name": "Saksoft"
    },

    {"uri": "https://balmer.finkraft.ai/auth/signin",
     "workspace_name": "Eastern regional office aicte kolkata",
     "table_name": "airline_recon_balmer",
     "db_workspace_name": "EASTERN REGIONAL OFFICE AICTE KOLKATA"
     },
    {
        "uri": "https://myyatra.finkraft.ai/auth/signin",
        "workspace_name": "Herbalife international india p ltd yatra",
        "table_name": "flight_recon_yatra",
        "db_workspace_name": "HERBALIFE INTERNATIONAL INDIA P LTD Yatra"
    },
    {
        "uri": "https://mmt.finkraft.ai/auth/signin",
        "workspace_name": "Kotak securities",
        "table_name": "airline_recon_mmt",
        "db_workspace_name": "Kotak Securities"
    },
    {
        "uri": "https://pyt.finkraft.ai/auth/signin",
        "workspace_name": "Blue heaven cosmetics",
        "table_name": "airline_recon_py",
        "db_workspace_name": "BLUE HEAVEN COSMETICS"
    },
    {
        "uri": "https://fcm.finkraft.ai/auth/signin",
        "workspace_name": "One ocean network",
        "table_name": "airline_recon_fcm",
        "db_workspace_name": "ONE OCEAN NETWORK"
    },
    {
        "uri": "https://bcd.finkraft.ai/auth/signin",
        "workspace_name": "Viewics",
        "table_name": "airline_recon_bcd",
        "db_workspace_name": "Viewics"
    },
    {
        "uri": "https://atpi.finkraft.ai/auth/signin",
        "workspace_name": "The great eastern shipping",
        "table_name": "airline_recon_atpi",
        "db_workspace_name": "The Great Eastern Shipping"
    }
]

run_set = os.environ.get("RUN_SET", "morning")
if run_set == "morning":
    portals = morning_workspaces
else:
    portals = afternoon_workspaces

login_username = os.environ.get("LOGIN_USERNAME")
login_password = os.environ.get("LOGIN_PASSWORD")
mongo_db_username = os.environ.get("MONGO_DB_USERNAME")
mongo_db_password = os.environ.get("MONGO_DB_PASSWORD")
pg_db_username = os.environ.get("PG_DB_USERNAME")
pg_db_password = os.environ.get("PG_DB_PASSWORD")

email_id = login_username
password = login_password


def initialize_driver(download_dir):
    print("Initializing the Chrome driver...")
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("--headless=new")  # Use "new" headless mode
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_experimental_option("prefs", {
        "download.default_directory": download_dir,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True
    })
    driver = webdriver.Chrome(options=chrome_options)
    driver.implicitly_wait(5)
    return driver


def login_and_select_workspace(driver, uri, workspace_name):
    wait = WebDriverWait(driver, 5)
    driver.get(uri)
    page_title = driver.title
    print(page_title)
    partner_portal_name = page_title.split("-")[0].strip()

    driver.find_element(By.XPATH, "//input[@placeholder='Email']").send_keys(email_id)
    driver.find_element(By.XPATH, '//button[@type="submit"]').click()
    driver.find_element(By.XPATH, "//input[@placeholder='Password']").send_keys(password)
    driver.find_element(By.CSS_SELECTOR, 'button[type="submit"]').click()
    print("Successfully logged in...")
    time.sleep(3)
    max_attempts = 4
    for attempt in range(max_attempts):
        try:
            workspace_dropdown = wait.until(
                EC.element_to_be_clickable((By.XPATH, "//div[contains(@class,'ant-dropdown-trigger')]")))
            workspace_dropdown.click()
            workspace_element = wait.until(
                EC.element_to_be_clickable((By.XPATH, f"//p[normalize-space()='{workspace_name}']")))
            try:
                workspace_element.click()
            except ElementClickInterceptedException:
                print("Element not clickable directly. Scrolling to it...")
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", workspace_element)
                time.sleep(1)  # Wait for scroll to complete
                workspace_element.click()
            time.sleep(1)
            menu_item = wait.until(EC.element_to_be_clickable((By.XPATH, '(//div[@class="MenuItem "])[2]')))
            menu_item.click()
            bulk_download = wait.until(
                EC.element_to_be_clickable((By.XPATH, "//span[normalize-space()='Download Report']")))
            bulk_download.click()

            # Generate a unique report name
            timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
            unique_id = str(uuid.uuid4())[:2]
            report_name = f"{workspace_name}_Report_{timestamp}_{unique_id}"

            report_name_input = wait.until(
                EC.presence_of_element_located((By.XPATH, "//input[contains(@placeholder,'Report Name')]")))
            report_name_input.clear()
            report_name_input.send_keys(report_name)

            try:
                download_button = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, '//span[normalize-space()="Create"]'))
                )
                download_button.click()
                ui_alert_shown_flag = True
                print("Alert shown")
            except Exception as e:
                ui_alert_shown_flag = False
                print("Alert not shown")

            return ui_alert_shown_flag, partner_portal_name, report_name, ""

        except Exception as e:
            print(f"Attempt {attempt + 1} failed: {str(e)}")
            if attempt < max_attempts - 1:
                print("Retrying workspace selection...")
                driver.refresh()
                time.sleep(2)
            else:
                print(f"Max attempts reached. Raising the error")
                return False, partner_portal_name, None, f"Workspace selection failed"

    return False, partner_portal_name, None, "Workspace selection failed after all retries"


def get_row_column_count_from_postgres(connection_params, table_name, workspace_name):
    """
    Connects to PostgreSQL and returns the count of rows filtered by workspace and the number of columns (minus 1).
    """
    max_retries = 3
    retry_delay = 2  # seconds
    conn = None
    cursor = None
    result = (None, None)

    for attempt in range(max_retries):
        try:
            conn = psycopg2.connect(
                host=connection_params['host'],
                port=connection_params['port'],
                database=connection_params['database'],
                user=connection_params['user'],
                password=connection_params['password']
            )
            cursor = conn.cursor()

            # Get row count
            row_query = sql.SQL('''
                SELECT COUNT(*)
                FROM {table}
                WHERE "Workspace" = %s
            ''').format(table=sql.Identifier(table_name))
            cursor.execute(row_query, (workspace_name,))
            row_count = cursor.fetchone()[0]

            # Get column count
            col_query = sql.SQL('''
                SELECT COUNT(*)
                FROM information_schema.columns
                WHERE table_name = %s
            ''')
            cursor.execute(col_query, (table_name,))
            col_count = cursor.fetchone()[0]
            col_count = col_count - 1

            cursor.close()
            conn.close()
            print(f" - Rows: {row_count}, Columns: {col_count}")
            return row_count, col_count

        except Exception as e:
            print(f"Error querying PostgreSQL (Attempt {attempt + 1}/{max_retries}): {e}")
            if attempt < max_retries - 1:
                print(f"Retrying in {retry_delay} seconds...")
                time.sleep(retry_delay)
                retry_delay *= 2  # Exponential backoff
            else:
                print("Max retries reached. Could not connect to PostgreSQL.")
                result = (None, None)
        finally:
            try:
                if cursor:
                    cursor.close()
                if conn:
                    conn.close()
            except Exception as e:
                print(f"Error closing PostgreSQL connection: {e}")
                return result
    return result


def remarks_to_mongo_db(partner_portal_name, workspace_name, db, current_timestamp, errormessage):
    error_collection = db["selenium-summary-excel-report"]
    data_to_insert = {
        "portalName": partner_portal_name,
        "workspaceName": workspace_name,
        "report_initialization_date_time": current_timestamp,
        "reportDownloadUIFlag": "--",
        "reportReceivedBackendFlag": "--",
        "totalColumnsInUI": "--",
        "totalRowsInUI": "--",
        "totalRowsInDB": "--",
        "totalRowsInExcel": "--",
        "totalColumnsInExcel": "--",
        "rowDifference": "--",
        "totalTime": "--",
        "testStatusWrtTime": "--",
        "testStatusWrtRow": "--",
        "remark": errormessage,
        "runId":run_id
    }
    error_collection.insert_one(data_to_insert)


def wait_for_report_completion(db, report_name, partner_portal_name, workspace_name):
    collection = db['recon_report']
    print(f"🔎 Searching in MongoDB for reportName: {report_name}...")
    doc = collection.find_one({"report_name": report_name})
    if not doc:
        current_timestamp = datetime.now().strftime('%Y-%m-%d')
        print("Document with reportName not found in MongoDB.")
        return "NO_REPORT_FOUND"  # Special return value for report not found case
        # Parse the createdAt field from MongoDB
    iso_date_val = doc["createdAt"]
    if isinstance(iso_date_val, str):
        utc_dt = parser.isoparse(iso_date_val)
    elif isinstance(iso_date_val, datetime):
        utc_dt = iso_date_val
    else:
        raise ValueError(f"Unexpected type for createdAt: {type(iso_date_val)}")

    if utc_dt.tzinfo is None:
        utc_dt = utc_dt.replace(tzinfo=pytz.UTC)
    ist_tz = pytz.timezone('Asia/Kolkata')
    ist_dt = utc_dt.astimezone(ist_tz)
    created_time_ts = ist_dt.timestamp()
    formatted_created_at_time = ist_dt.strftime('%d-%b-%Y %I:%M %p')
    print(f"🕒 Created At (IST): {formatted_created_at_time}")

    timeout = (60 * 30)  # 30 minutes
    poll_interval = 10  # every 10 seconds
    start_time = time.time()
    while time.time() - start_time < timeout:
        doc = collection.find_one({"report_name": report_name})
        if doc and doc["status"] == "COMPLETED":
            completed_time = time.time()
            print("Status changed to COMPLETED at", time.ctime(completed_time))
            total_time = int(completed_time - created_time_ts)
            print(f"Total time taken: {total_time} seconds")
            total_files = doc.get("total_record")
            if total_files is None:
                print("⚠️ 'totalFiles' field not found in MongoDB document.")
                total_files = 0
            else:
                print(f"🗃️ Total number of rows items in excel to be downloaded as per DB: {total_files}")
            file_hash = doc.get("filehash")
            return {
                "total_files": total_files,
                "file_hash": file_hash,
                "total_time": total_time,
                "formatted_created_at_time": formatted_created_at_time
            }
        print("⌛ Still PENDING... checking again in 10 seconds")
        time.sleep(poll_interval)

    print("Timeout: Report did not complete in estimated minutes.")
    print("❌Error Report inserted successfully into DB")
    return "TIMEOUT"  # Special return value for timeout case


def download_and_verify_invoices(driver, file_hash, total_row_db, total_time, formatted_created_at_time,
                                 partner_portal_name, workspace_name, ui_alert_shown_flag, db, download_dir,
                                 ui_row_count, ui_column_count):
    if not file_hash:
        print("❌ fileHash not found in MongoDB.")
        return False, "fileHash not found in MongoDB."
    file_hash_flag = True
    print("✅ fileHash is present. Flag:", file_hash_flag)
    download_url = f"https://files.finkraft.ai/invoice-{file_hash}"
    print(f"🌐 Navigating to file download URL")
    driver.get(download_url)
    time.sleep(3)
    try:
        download_btn_1 = driver.find_element(By.XPATH, "//a[@id='downloadLink']")
        download_btn_1.click()
        time.sleep(3)
        print("⬇️ Invoice download links clicked successfully.")
    except Exception as e:
        print(f"❌ Error clicking download buttons: {e}")
        return False, f"Error clicking download buttons: {str(e)}"

    # Wait for the .xlsx file to appear in the download directory
    xlsx_file_path = None
    print("⏳ Waiting for .xlsx file to appear in Downloads...")
    timeout = time.time() + 70  # wait up to 70 seconds
    while time.time() < timeout:
        xlsx_files = glob.glob(os.path.join(download_dir, "*.xlsx"))
        if xlsx_files:
            xlsx_file_path = max(xlsx_files, key=os.path.getctime)  # get latest xlsx
            print(f"✅ Found xlsx file: {xlsx_file_path}")
            break
        time.sleep(2)

    if not xlsx_file_path:
        print("❌ XLSX file not downloaded.")
        return False, "XLSX file not downloaded."

    # Open the xlsx file and check row and column count
    wb = openpyxl.load_workbook(xlsx_file_path)
    ws = wb.active
    excel_row_count = ws.max_row - 1  # Exclude header
    excel_col_count = ws.max_column
    print(f"📄 XLSX file row count: {excel_row_count}, column count: {excel_col_count}")

    # Delete the xlsx file after processing
    try:
        os.remove(xlsx_file_path)
        print("🧹 Deleted downloaded XLSX file.")
    except Exception as e:
        print(f"⚠️ Error deleting XLSX file: {e}")

    print("🧮 Verifying report download count consistency...")
    print(f" - 📦 Total rows item reported in MongoDB:  {total_row_db}")
    print(f" - 📄 Total Report invoices extracted from the downloaded XlSX: {excel_row_count}")

    if excel_row_count == 0:
        print("❌ No rows found in downloaded file. Cannot calculate per-invoice time.")
        return False, "No rows found in downloaded file. Cannot calculate per-invoice time."

    row_diff = total_row_db - excel_row_count
    test_status_wrt_time = "PASS" if total_time <= (30 * 60) else "FAIL"
    test_status_wrt_row = "PASS" if row_diff == 0 else "FAIL"
    total_time_in_min = round(total_time / 60, 2)

    summary_collection = db["selenium-summary-excel-report"]
    data_to_insert = {
        "portalName": partner_portal_name,
        "workspaceName": workspace_name,
        "report_initialization_date_time": formatted_created_at_time,
        "reportDownloadUIFlag": ui_alert_shown_flag,
        "reportReceivedBackendFlag": file_hash_flag,
        "totalColumnsInUI": ui_column_count,
        "totalRowsInUI": ui_row_count,
        "totalRowsInDB": total_row_db,
        "totalRowsInExcel": excel_row_count,
        "totalColumnsInExcel": excel_col_count,
        "rowDifference": row_diff,
        "totalTime": total_time_in_min,
        "testStatusWrtTime": test_status_wrt_time,
        "testStatusWrtRow": test_status_wrt_row,
        "remark": "",
        "runId":run_id
    }
    summary_collection.insert_one(data_to_insert)
    print("✅ Report inserted successfully into DB")

    if row_diff == 0:
        print("🟩 Report count matches exactly between DB and downloaded folder.")
    elif row_diff > 0:
        print(f"🟨 {row_diff} invoice(s) reported in DB but missing in folder.")
    else:
        print(f"🟥 {abs(row_diff)} extra invoice(s) found in folder not reported in DB.")

    if total_time > 60 * 30:
        print("❌ Test Failed: More than 2 seconds per invoice")
    else:
        print("✅ Test Passed: Invoice download time is within limits")
    return True,""

def main():

    for portal in portals:
        retry_count = 0
        max_retries = 1  # Only retry once
        while retry_count <= max_retries:
            connection_string = (f"mongodb://{mongo_db_username}:{mongo_db_password}"
                                 "@mongodb.internal.finkraftai.com/admin?"
                                 "directConnection=true&serverSelectionTimeoutMS=20000&appName=mongosh+2.2.3")
            client = MongoClient(connection_string)
            db = client['gstservice']
            print(f"Starting automation for portal:{portal['uri']} and workspace:{portal['workspace_name']}")
            download_dir = tempfile.mkdtemp(prefix="downloads_")
            driver = initialize_driver(download_dir)
            try:
                ui_alert_shown_flag, partner_portal_name, report_name, error_message = login_and_select_workspace(
                    driver,
                    portal["uri"],
                    portal["workspace_name"]
                )

                if not ui_alert_shown_flag:
                    print("Workspace selection failed after all retries. Moving to next portal.")
                    current_timestamp = datetime.now().strftime('%Y-%m-%d')
                    remarks_to_mongo_db(partner_portal_name, portal['workspace_name'], db, current_timestamp,
                                        error_message)
                    break

                pg_connection_params = {
                    "host": "postgresql.internal.finkraftai.com",
                    "port": 5432,
                    "database": "airlines_db",
                    "user": pg_db_username,
                    "password": pg_db_password
                }
                table_name = portal['table_name']
                workspace_name = portal['db_workspace_name']

                row_count, col_count = get_row_column_count_from_postgres(
                    pg_connection_params,
                    table_name,
                    workspace_name
                )
                if row_count is None or col_count is None:
                    print("❌ Failed to get row and column count from PostgreSQL. Skipping further processing.")
                    current_timestamp = datetime.now().strftime('%Y-%m-%d')
                    remarks_to_mongo_db(partner_portal_name, portal['workspace_name'], db, current_timestamp,
                                        "Failed to get row and column count from PostgreSQL")
                    break

                report_info = wait_for_report_completion(db, report_name, partner_portal_name,
                                                         portal["workspace_name"])

                if report_info == "NO_REPORT_FOUND":
                    print("Report not found in MongoDB. Will retry...")
                    if retry_count == 0:
                        retry_count += 1
                        continue
                    current_timestamp = datetime.now().strftime('%Y-%m-%d')
                    remarks_to_mongo_db(partner_portal_name, portal['workspace_name'], db, current_timestamp,
                                        "Report not found in MongoDB after retry")
                    break
                elif report_info == "TIMEOUT":
                    print("Report did not complete within 30 minutes. Moving to next portal.")
                    current_timestamp = datetime.now().strftime('%Y-%m-%d')
                    remarks_to_mongo_db(partner_portal_name, portal['workspace_name'], db, current_timestamp,
                                        "Report did not complete within 30 minutes")
                    break  # No retry for timeout, move to next portal
                elif report_info is None:
                    print("Unexpected error in report completion. Moving to next portal.")
                    current_timestamp = datetime.now().strftime('%Y-%m-%d')
                    remarks_to_mongo_db(partner_portal_name, portal['workspace_name'], db, current_timestamp,
                                        "Unexpected error in report completion")
                    break

                download_success, error_message = download_and_verify_invoices(
                    driver,
                    report_info["file_hash"],
                    report_info["total_files"],
                    report_info["total_time"],
                    report_info["formatted_created_at_time"],
                    partner_portal_name,
                    portal['workspace_name'],
                    ui_alert_shown_flag,
                    db, download_dir, row_count, col_count
                )

                if download_success is False:
                    print(f"Download verification failed: {error_message}")
                    if retry_count == 0:
                        retry_count += 1
                        continue
                    current_timestamp = datetime.now().strftime('%Y-%m-%d')
                    remarks_to_mongo_db(partner_portal_name, portal['workspace_name'], db, current_timestamp,
                                        error_message)
                    break

            except Exception as e:
                print(f"⚠️ Error processing portal {portal['uri']}: {e}")
                try:
                    current_timestamp = datetime.now().strftime('%Y-%m-%d')
                    subdomain = portal['uri'].split('//')[-1].split('/')[0].replace('.finkraft.ai', '')
                    max_length = 50
                    short_exception = str(e)[:max_length] + ("..." if len(str(e)) > max_length else "")
                    remarks_to_mongo_db(subdomain, portal['workspace_name'], db, current_timestamp,
                                        f" Exception during processing: {short_exception}")
                except Exception as log_e:
                    print(f"❌ Error while attempting to log exception to MongoDB: {log_e}")

            finally:
                if driver:
                    driver.quit()

            print(f"Completed automation for portal: {portal['uri']}\n\n")
            break


if __name__ == "__main__":
    main()